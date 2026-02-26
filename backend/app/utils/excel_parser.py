"""
Excel import utility: reads an .xlsx file and returns list of dicts,
mapping column headers (or positional indices) to item fields.
"""
from typing import List, Dict, Any, Optional
import openpyxl
from io import BytesIO


COLUMN_ALIASES = {
    # item_code
    "code": "item_code", "item code": "item_code", "itemcode": "item_code",
    # description
    "desc": "description", "item description": "description", "name": "description",
    # quantity
    "qty": "quantity", "qnty": "quantity",
    # unit
    "uom": "unit",
    # brand / model (installation)
    "make": "brand", "manufacturer": "brand",
    # part_number (installation)
    "part no": "part_number", "part#": "part_number",
    # location_zone (installation)
    "zone": "location_zone", "area": "location_zone",
    # floor_level (installation)
    "floor": "floor_level", "level": "floor_level",
    # asset_tag (maintenance)
    "asset": "asset_tag", "tag": "asset_tag",
    # task_description (maintenance)
    "task": "task_description",
    # task_name (programming_handover)
    "task name": "task_name",
    # condition (survey)
    "cond": "condition",
    # remarks
    "remark": "remarks", "comment": "remarks", "comments": "remarks",
    # file_name (handover)
    "document": "file_name", "file": "file_name",
}


def normalize_header(h: str) -> str:
    return h.strip().lower()


def parse_excel(content: bytes, column_map: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
    """
    Parse an Excel file from bytes.
    column_map: optional dict from Excel column header -> field name override.
    Returns list of dicts with raw row data.
    """
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # First row = headers
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

    # Build field mapping: excel_header -> model_field
    field_map: Dict[str, str] = {}
    for h in headers:
        norm = normalize_header(h)
        if column_map and h in column_map:
            field_map[h] = column_map[h]
        elif norm in COLUMN_ALIASES:
            field_map[h] = COLUMN_ALIASES[norm]
        else:
            field_map[h] = norm.replace(" ", "_")

    items = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue  # skip empty rows

        record: Dict[str, Any] = {}
        for col_idx, header in enumerate(headers):
            value = row[col_idx] if col_idx < len(row) else None
            field_name = field_map.get(header, header)
            if value is not None:
                record[field_name] = value

        if record:
            items.append(record)

    return items


def get_excel_headers(content: bytes) -> List[str]:
    """Return the header row from an Excel file."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    return [str(h).strip() if h is not None else "" for h in first_row]
